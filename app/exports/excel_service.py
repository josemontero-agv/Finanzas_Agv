# -*- coding: utf-8 -*-
"""
Servicio de Exportación a Excel.

Genera archivos Excel a partir de datos de reportes.
"""

from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExportService:
    """
    Servicio para exportar datos a archivos Excel con formato profesional.
    """
    
    # Estilos predefinidos
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    CELL_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def export_collections_report(data, filename="reporte_cobranzas.xlsx"):
        """
        Exporta reporte de cobranzas a Excel con todas las columnas.
        Formato simplificado y profesional similar al reporte de tesorería.
        
        Args:
            data (list): Lista de diccionarios con datos del reporte
            filename (str): Nombre del archivo a generar
        
        Returns:
            BytesIO: Buffer con el archivo Excel generado
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CxC - Cuenta 12"
        
        # Definir todas las columnas (29 columnas - incluye campos calculados)
        columns = [
            ('invoice_date', 'Fecha Factura'),
            ('date', 'Fecha Contabilización'),
            ('l10n_latam_document_type_id', 'Tipo Documento'),
            ('move_name', 'Número Documento'),
            ('l10n_latam_boe_number', 'Letra'),
            ('invoice_origin', 'Origen'),
            ('account_id/code', 'Cuenta'),
            ('account_id/name', 'Nombre Cuenta'),
            ('patner_id/vat', 'RUC/DNI'),
            ('patner_id', 'Cliente'),
            ('currency_id', 'Moneda'),
            ('amount_currency', 'Total Moneda Origen'),
            ('amount_residual_with_retention', 'Adeudado'),
            ('amount_residual_signed', 'Adeudado S/.'),
            ('date_maturity', 'Fecha Vencimiento'),
            ('dias_vencido', 'Días Vencido'),
            ('estado_deuda', 'Estado'),
            ('antiguedad', 'Antigüedad'),
            ('ref', 'Referencia'),
            ('invoice_payment_term_id', 'Condición Pago'),
            ('name', 'Descripción'),
            ('move_id/invoice_user_id', 'Vendedor'),
            ('patner_id/state_id', 'Provincia'),
            ('patner_id/l10n_pe_district', 'Distrito'),
            ('patner_id/country_code', 'Código País'),
            ('patner_id/country_id', 'País'),
            ('partner_groups', 'Grupos'),
            ('sub_channel_id', 'Sub Canal'),
            ('move_id/sales_channel_id', 'Canal de Venta'),
            ('move_id/sales_type_id', 'Tipo de Venta'),
        ]
        
        # Escribir encabezados
        for col_num, (key, header) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = ExcelExportService.HEADER_FILL
            cell.font = ExcelExportService.HEADER_FONT
            cell.alignment = ExcelExportService.HEADER_ALIGNMENT
            cell.border = ExcelExportService.CELL_BORDER
        
        # Escribir datos
        for row_num, record in enumerate(data, 2):
            for col_num, (key, header) in enumerate(columns, 1):
                value = record.get(key, '')
                
                # Convertir valores Many2One (listas) a string
                if isinstance(value, (list, tuple)) and len(value) >= 2:
                    value = str(value[1])  # Extraer el nombre
                elif isinstance(value, (list, tuple)):
                    value = str(value[0]) if value else ''
                
                # Convertir None a cadena vacía
                if value is None:
                    value = ''
                
                # Formatear valores numéricos
                if key in ['amount_currency', 'amount_residual_with_retention', 'amount_residual_signed', 'amount_total', 'dias_vencido']:
                    try:
                        value = float(value) if value else 0
                        if key == 'dias_vencido':
                            value = int(value)
                    except:
                        value = 0
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = ExcelExportService.CELL_BORDER
                
                # Formato especial para números
                if key in ['amount_currency', 'amount_residual_with_retention', 'amount_residual_signed', 'amount_total']:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif key == 'dias_vencido':
                    cell.alignment = Alignment(horizontal='center')
                    # Resaltar en rojo si está vencido
                    if value > 0:
                        cell.font = Font(color="FF0000", bold=True)
                elif key == 'estado_deuda':
                    cell.alignment = Alignment(horizontal='center')
                    # Color de fondo según estado
                    if value == 'VENCIDO':
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # Ajustar anchos de columna según contenido
        column_widths = {
            1: 12,  # Fecha Factura
            2: 14,  # Fecha Contabilización
            3: 14,  # Tipo Documento
            4: 16,  # Número Documento
            5: 12,  # Origen
            6: 10,  # Cuenta
            7: 25,  # Nombre Cuenta
            8: 12,  # RUC/DNI
            9: 30,  # Cliente
            10: 10, # Moneda
            11: 18, # Total Moneda Origen
            12: 18, # Adeudado
            13: 18, # Adeudado S/.
            14: 14, # Fecha Vencimiento
            15: 12, # Días Vencido
            16: 12, # Estado
            17: 20, # Antigüedad
            18: 14, # Referencia
            19: 18, # Condición Pago
            20: 30, # Descripción
            21: 20, # Vendedor
            22: 18, # Provincia
            23: 18, # Distrito
            24: 12, # Código País
            25: 18, # País
            26: 25, # Grupos
            27: 18, # Sub Canal
            28: 20, # Canal de Venta
            29: 18, # Tipo de Venta
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def export_treasury_report(data, filename="reporte_tesoreria.xlsx"):
        """
        Exporta reporte de tesorería (CxP) a Excel con todos los campos expandidos.
        Formato similar al reporte de collections.
        
        Args:
            data (list): Lista de diccionarios con datos del reporte
            filename (str): Nombre del archivo a generar
        
        Returns:
            BytesIO: Buffer con el archivo Excel generado
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CxP - Cuenta 42"
        
        # Definir columnas expandidas (28 campos)
        columns = [
            ('invoice_date', 'Fecha Factura'),
            ('date', 'Fecha Contabilización'),
            ('l10n_latam_document_type_id', 'Tipo Documento'),
            ('move_name', 'Número Documento'),
            ('ref', 'Referencia'),
            ('invoice_origin', 'Origen'),
            ('account_code', 'Cuenta'),
            ('account_name', 'Nombre Cuenta'),
            ('supplier_vat', 'RUC Proveedor'),
            ('supplier_name', 'Proveedor'),
            ('supplier_country_code', 'Código País'),
            ('supplier_country', 'País'),
            ('supplier_state', 'Provincia'),
            ('supplier_city', 'Ciudad'),
            ('supplier_phone', 'Teléfono'),
            ('supplier_email', 'Email'),
            ('currency_id', 'Moneda'),
            ('amount_total_in_currency_signed', 'Total Origen'),
            ('amount_residual_with_retention', 'Adeudado Origen'),
            ('amount_total_signed', 'Total S/.'),
            ('invoice_date_due', 'Fecha Vencimiento'),
            ('dias_vencido', 'Días Vencido'),
            ('estado_deuda', 'Estado'),
            ('antiguedad', 'Antigüedad'),
            ('invoice_payment_term_id', 'Condición Pago'),
            ('payment_state', 'Estado Pago'),
            ('state', 'Estado Factura'),
            ('invoice_user_id', 'Usuario Responsable'),
            ('name', 'Descripción'),
        ]
        
        # Escribir encabezados
        for col_num, (key, header) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = ExcelExportService.HEADER_FILL
            cell.font = ExcelExportService.HEADER_FONT
            cell.alignment = ExcelExportService.HEADER_ALIGNMENT
            cell.border = ExcelExportService.CELL_BORDER
        
        # Escribir datos
        for row_num, record in enumerate(data, 2):
            for col_num, (key, header) in enumerate(columns, 1):
                value = record.get(key, '')
                
                # Convertir valores Many2One (listas) a string
                if isinstance(value, (list, tuple)) and len(value) >= 2:
                    value = str(value[1])  # Extraer el nombre
                elif isinstance(value, (list, tuple)):
                    value = str(value[0]) if value else ''
                
                # Convertir None a cadena vacía
                if value is None:
                    value = ''
                
                # Formatear valores numéricos
                if key in ['amount_total_in_currency_signed', 'amount_residual_with_retention', 'amount_total_signed', 'dias_vencido']:
                    try:
                        value = float(value) if value else 0
                        if key == 'dias_vencido':
                            value = int(value)
                    except:
                        value = 0
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = ExcelExportService.CELL_BORDER
                
                # Formato especial para números
                if key in ['amount_total_in_currency_signed', 'amount_residual_with_retention', 'amount_total_signed']:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif key == 'dias_vencido':
                    cell.alignment = Alignment(horizontal='center')
                    # Resaltar en rojo si está vencido
                    if value > 0:
                        cell.font = Font(color="FF0000", bold=True)
                elif key == 'estado_deuda':
                    cell.alignment = Alignment(horizontal='center')
                    # Color de fondo según estado
                    if value == 'VENCIDO':
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # Ajustar anchos de columna
        column_widths = [12, 14, 14, 16, 14, 12, 10, 25, 14, 30, 12, 18, 18, 18, 16, 25, 10, 18, 18, 18, 14, 12, 12, 20, 18, 14, 14, 20, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer

